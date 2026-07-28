"""
Экспорт/импорт Excel для журналов "Учёт (автомойка)" и "Химчистка".

Формат — простой, один лист, заголовок + строки, с колонками ровно как в
ответах API (WalkInOrderOut / DryCleaningOrderOut). Экспортировал → открыл в
Excel → поправил/добавил строки → импортировал обратно — тот же формат.

Это НЕ то же самое, что app/scripts/import_uchet.py — тот скрипт разбирает
твой исходный многолистовой "Учёт.xlsx" (историческая миграция один раз).
Этот модуль — для повседневного экспорта/импорта уже нормализованных данных
через админку.
"""
import io
import re
from datetime import date, datetime
from typing import List

import openpyxl
from openpyxl.utils import get_column_letter

WALKIN_HEADERS = [
    "Дата", "Время (текст)", "Услуга", "Доп. услуга",
    "Марка авто", "Сумма", "Контакт", "ID сотрудника",
]

DRYCLEANING_HEADERS = [
    "Дата", "Марка авто", "Работы", "Телефон",
    "Сумма", "З/п мастера", "ID сотрудника",
]


def _autofit(ws):
    for i, _ in enumerate(ws[1], start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def export_walkin_xlsx(orders: List) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Учёт"
    ws.append(WALKIN_HEADERS)
    for o in orders:
        ws.append([
            o.order_date.isoformat() if o.order_date else "",
            o.time_note or "",
            o.service_name_raw or "",
            o.extra_service or "",
            o.car_model or "",
            o.amount or 0,
            o.contact_name or "",
            o.employee_id or "",
        ])
    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_drycleaning_xlsx(orders: List) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Химчистка"
    ws.append(DRYCLEANING_HEADERS)
    for o in orders:
        ws.append([
            o.order_date.isoformat() if o.order_date else "",
            o.car_model or "",
            o.works_description or "",
            o.phone or "",
            o.amount or 0,
            o.employee_payout or "",
            o.employee_id or "",
        ])
    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    # пробуем несколько распространённых форматов, а не только ISO
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"не удалось распознать дату {s!r} (ожидается формат ГГГГ-ММ-ДД или ДД.ММ.ГГГГ)")


# Псевдонимы заголовков — колонки ищутся по названию, порядок в файле не важен.
# Ключ — как выглядит заголовок в файле (без учёта регистра/пробелов), значение — наше поле.
_WALKIN_ALIASES = {
    "дата": "order_date",
    "время (текст)": "time_note", "время": "time_note",
    "услуга": "service_name_raw", "наименование услуги": "service_name_raw",
    "доп. услуга": "extra_service", "доп услуга": "extra_service", "доп. услуги": "extra_service",
    "марка авто": "car_model", "авто": "car_model", "марка": "car_model", "марка автомобиля": "car_model", "автомобиль": "car_model",
    "сумма": "amount", "стоимость": "amount",
    "контакт": "contact_name",
    "id сотрудника": "employee_id",
}
_WALKIN_REQUIRED = ["service_name_raw", "amount"]

class HeaderMismatch(ValueError):
    """Заголовки файла не похожи на наш экспортный формат — сигнал пробовать
    родной многолистовой формат (Учёт.xlsx/Химчистка.xlsx)."""
    pass


_DRYCLEANING_ALIASES = {
    "дата": "order_date",
    "марка авто": "car_model", "авто": "car_model", "марка": "car_model", "автомобиль": "car_model",
    "работы": "works_description", "услуга": "works_description",
    "телефон": "phone", "контакт": "phone",
    "сумма": "amount", "стоимость": "amount",
    "з/п мастера": "employee_payout", "зп мастера": "employee_payout", "зарплата": "employee_payout",
    "з/п": "employee_payout", "зп": "employee_payout",
    "id сотрудника": "employee_id",
}
_DRYCLEANING_REQUIRED = ["car_model", "works_description", "amount"]


def _build_column_map(header_row, aliases: dict, required: list, required_labels: dict):
    """По строке заголовков строит {индекс_колонки: имя_поля}, проверяя, что все
    обязательные поля найдены. Если что-то не найдено — кидает понятную ошибку
    со списком того, что реально было в заголовке файла."""
    col_map = {}
    for i, cell in enumerate(header_row):
        if not cell:
            continue
        key = str(cell).strip().lower()
        field = aliases.get(key)
        if field:
            col_map[i] = field

    found_fields = set(col_map.values())
    missing = [f for f in required if f not in found_fields]
    if missing:
        missing_labels = ", ".join(required_labels[f] for f in missing)
        seen_headers = ", ".join(str(c) for c in header_row if c) or "(пусто)"
        raise HeaderMismatch(
            f"в файле не найдены обязательные колонки: {missing_labels}. "
            f"Колонки, которые увидел в первой строке: {seen_headers}"
        )
    return col_map


def parse_walkin_xlsx(file_bytes: bytes) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("файл пустой")

    col_map = _build_column_map(
        header_row, _WALKIN_ALIASES, _WALKIN_REQUIRED,
        {"service_name_raw": "Услуга", "amount": "Сумма"},
    )

    rows = []
    for line_num, row in enumerate(rows_iter, start=2):
        values = {field: row[i] if i < len(row) else None for i, field in col_map.items()}
        service = values.get("service_name_raw")
        amount = values.get("amount")
        if not service and amount in (None, ""):
            continue  # пустая строка — пропускаем молча
        if not service or amount in (None, ""):
            raise ValueError(f"строка {line_num}: не заполнены обязательные поля (Услуга/Сумма)")
        try:
            amount_val = float(amount)
        except (TypeError, ValueError):
            raise ValueError(f"строка {line_num}: в колонке 'Сумма' не число ({amount!r})")

        employee_id = values.get("employee_id")
        rows.append({
            "order_date": _parse_date(values.get("order_date")) if values.get("order_date") else date.today(),
            "time_note": str(values["time_note"]) if values.get("time_note") else None,
            "service_name_raw": str(service),
            "extra_service": str(values["extra_service"]) if values.get("extra_service") else None,
            "car_model": str(values["car_model"]) if values.get("car_model") else None,
            "amount": amount_val,
            "contact_name": str(values["contact_name"]) if values.get("contact_name") else None,
            "employee_id": int(employee_id) if employee_id not in (None, "") else None,
        })
    return rows


def parse_drycleaning_xlsx(file_bytes: bytes) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("файл пустой")

    col_map = _build_column_map(
        header_row, _DRYCLEANING_ALIASES, _DRYCLEANING_REQUIRED,
        {"car_model": "Марка авто", "works_description": "Работы", "amount": "Сумма"},
    )

    rows = []
    for line_num, row in enumerate(rows_iter, start=2):
        values = {field: row[i] if i < len(row) else None for i, field in col_map.items()}
        car = values.get("car_model")
        works = values.get("works_description")
        amount = values.get("amount")
        if not car and not works and amount in (None, ""):
            continue
        if not car or amount in (None, ""):
            continue  # неполная историческая строка (нет марки или суммы вообще) — молча пропускаем
        try:
            amount_val = float(amount)
        except (TypeError, ValueError):
            raise ValueError(f"строка {line_num}: в колонке 'Сумма' указано не число ({amount!r})")

        payout = values.get("employee_payout")
        employee_id = values.get("employee_id")
        rows.append({
            "order_date": _parse_date(values.get("order_date")) if values.get("order_date") else date.today(),
            "car_model": str(car),
            "works_description": str(works) if works else "Без описания",
            "phone": str(values["phone"]) if values.get("phone") else None,
            "amount": amount_val,
            "employee_payout": float(payout) if payout not in (None, "") else None,
            "employee_id": int(employee_id) if employee_id not in (None, "") else None,
        })
    return rows


# ---------------------------------------------------------------------------
# "Родной" формат — исходный многолистовой "Учёт.xlsx" (день-блоками, без
# единого заголовка). Распознаём по провалу поиска заголовков в чистом формате.
# ---------------------------------------------------------------------------

_SERVICE_NORMALIZE = [
    (re.compile(r"компл", re.I), "Комплекс"),
    (re.compile(r"эксп?р[еэ]с+", re.I), "Экспресс"),
    (re.compile(r"облив", re.I), "Облив"),
    (re.compile(r"салон", re.I), "Химчистка салона"),
]


def _normalize_service_name(raw: str) -> str:
    if not raw:
        return "Без названия"
    for pattern, canonical in _SERVICE_NORMALIZE:
        if pattern.search(raw):
            return canonical
    return raw.strip()


def parse_walkin_raw_multisheet(file_bytes: bytes) -> List[dict]:
    """Разбирает исходный 'Учёт.xlsx': один лист на месяц, внутри — блоки по
    дням (строка-дата, затем несколько строк-заказов, иногда 'Итого:')."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    rows_out = []
    current_date = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            time_note, num, marka, service_raw, extra, amount, contact = (list(row) + [None] * 7)[:7]

            if isinstance(num, datetime):
                current_date = num.date()
                continue
            if isinstance(service_raw, datetime):
                current_date = service_raw.date()
                continue

            if not service_raw or not isinstance(amount, (int, float)):
                continue  # заголовки/итоги/пустые строки — пропускаем молча
            if current_date is None:
                continue

            rows_out.append({
                "order_date": current_date,
                "time_note": str(time_note) if time_note else None,
                "service_name_raw": _normalize_service_name(str(service_raw)),
                "extra_service": str(extra) if extra else None,
                "car_model": str(marka) if marka else None,
                "amount": float(amount),
                "contact_name": str(contact) if contact else None,
                "employee_id": None,
            })
    return rows_out


def parse_walkin_any(file_bytes: bytes) -> List[dict]:
    """Сначала пробуем наш чистый экспортный формат (по заголовкам). Если
    заголовки не похожи — пробуем родной многолистовой 'Учёт.xlsx'."""
    try:
        return parse_walkin_xlsx(file_bytes)
    except HeaderMismatch:
        rows = parse_walkin_raw_multisheet(file_bytes)
        if not rows:
            raise ValueError(
                "не удалось распознать файл ни в одном из известных форматов "
                "(ни как экспорт из админки, ни как исходный 'Учёт.xlsx')"
            )
        return rows
