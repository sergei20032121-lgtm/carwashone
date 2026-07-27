"""
Импорт исторического журнала заказов из "Учёт.xlsx" в таблицу walk_in_orders.

Формат файла (как прислал): один лист на месяц, внутри — блоки по дням:
  строка с датой | несколько строк вида (№, Марка авто, Услуга, Доп.услуга, Сумма, Контакт) | строка "Итого:"

Запуск:
    python -m app.scripts.import_uchet "путь/к/Учет.xlsx"

Названия услуг в файле разнобойные ("Комплексная", "комплексная", "Комплесная" и т.п.) —
скрипт нормализует их в несколько известных категорий и, если получается,
привязывает к Service из каталога (по needle-совпадению с названием).
Если совпадения нет — просто сохраняет как есть в service_name_raw, это не потеряется.
"""
import sys
import re
from datetime import datetime

import openpyxl

from app.database import SessionLocal
from app.models import WalkInOrder, Service

# нормализация "грязных" названий услуг из журнала -> к чему привести
NORMALIZE = [
    (re.compile(r"компл", re.I), "Комплекс"),
    (re.compile(r"эксп?р[еэ]с+", re.I), "Экспресс"),
    (re.compile(r"облив", re.I), "Облив"),
    (re.compile(r"салон", re.I), "Химчистка салона"),
]


def normalize_service_name(raw: str) -> str:
    if not raw:
        return "Без названия"
    for pattern, canonical in NORMALIZE:
        if pattern.search(raw):
            return canonical
    return raw.strip()


def find_service_id(db, canonical_name: str):
    svc = db.query(Service).filter(Service.name.ilike(f"%{canonical_name}%")).first()
    return svc.id if svc else None


def import_file(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    db = SessionLocal()
    imported = 0
    skipped = 0
    current_date = None

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                time_note, num, marka, service_raw, extra, amount, contact = (list(row) + [None] * 7)[:7]

                # строка-дата: только 4-й элемент — datetime, остальное пусто
                if isinstance(num, datetime):
                    current_date = num.date()
                    continue
                if isinstance(service_raw, datetime):
                    current_date = service_raw.date()
                    continue

                # пропускаем заголовки/итоги/пустые строки
                if not service_raw or not isinstance(amount, (int, float)):
                    skipped += 1
                    continue
                if current_date is None:
                    skipped += 1
                    continue

                canonical = normalize_service_name(str(service_raw))
                service_id = find_service_id(db, canonical)

                db.add(WalkInOrder(
                    order_date=current_date,
                    time_note=str(time_note) if time_note else None,
                    service_id=service_id,
                    service_name_raw=canonical,
                    extra_service=str(extra) if extra else None,
                    car_model=str(marka) if marka else None,
                    amount=float(amount),
                    contact_name=str(contact) if contact else None,
                ))
                imported += 1

        db.commit()
        print(f"Импортировано записей: {imported}, пропущено (итоги/пустые строки): {skipped}")
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python -m app.scripts.import_uchet путь/к/Учет.xlsx")
        sys.exit(1)
    import_file(sys.argv[1])
